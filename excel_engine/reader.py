"""
Excel reader for loading IS Claim data from Sheet1.
Uses openpyxl for precise cell-level access with column index mapping.
"""

import openpyxl
from datetime import datetime, date
from utils.constants import ClaimCol


def load_workbook(file_path: str):
    """
    Load the Excel workbook and return the Sheet1 worksheet.

    Returns:
        ws: The Sheet1 worksheet object.
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except PermissionError:
        raise PermissionError(
            f"Cannot open '{file_path}' — file is locked. "
            f"Please CLOSE the Excel file and try again."
        )

    if "Sheet1" not in wb.sheetnames:
        raise ValueError("Excel file must contain a 'Sheet1' sheet")

    return wb["Sheet1"]


def get_total_rows(ws) -> int:
    """Get the total number of data rows (excluding header)."""
    max_row = ws.max_row
    # Find actual last row with data in the Loan App No column
    for row in range(max_row, 1, -1):
        if ws.cell(row=row, column=ClaimCol.LOAN_APP_NO).value is not None:
            return row
    return 1  # Only header


def _format_date(val) -> str:
    """Format a datetime/date value to DD-MM-YYYY string for the portal."""
    if isinstance(val, datetime):
        return val.strftime("%d-%m-%Y")
    if isinstance(val, date):
        return val.strftime("%d-%m-%Y")
    # Already a string — pass through
    return str(val).strip()


def _cell_str(ws, row, col) -> str:
    """Get cell value as a stripped string, or empty string if None."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip()


def _cell_numeric(ws, row, col) -> str:
    """Get cell value as a clean numeric string (remove duplicate dots, etc.)."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        # Format cleanly — avoid trailing .0 for integers
        if isinstance(val, float) and val == int(val):
            return str(int(val))
        return str(val)
    # String value — clean up
    s = str(val).strip()
    # Remove duplicate dots (e.g., "78..04" → "78.04")
    while ".." in s:
        s = s.replace("..", ".")
    return s


def _cell_val(ws, row, col):
    """Get raw cell value."""
    return ws.cell(row=row, column=col).value


def read_claim_row(ws, row: int) -> dict:
    """
    Read a single row from Sheet1 and return as a dictionary.
    Row is 1-based (row 2 = first data row).
    """
    return {
        "sr": _cell_val(ws, row, ClaimCol.SR),
        "loan_app_no": _cell_str(ws, row, ClaimCol.LOAN_APP_NO),
        "first_disbursal_date": _format_date(_cell_val(ws, row, ClaimCol.FIRST_DISBURSAL_DATE)) if _cell_val(ws, row, ClaimCol.FIRST_DISBURSAL_DATE) else "",
        "rollover_date": _format_date(_cell_val(ws, row, ClaimCol.ROLLOVER_DATE)) if _cell_val(ws, row, ClaimCol.ROLLOVER_DATE) else "",
        "max_withdrawal": _cell_numeric(ws, row, ClaimCol.MAX_WITHDRAWAL),
        "applicable_is": _cell_numeric(ws, row, ClaimCol.APPLICABLE_IS),
        "claim_id": _cell_str(ws, row, ClaimCol.CLAIM_ID),
        "season": _cell_str(ws, row, ClaimCol.SEASON),
        "account_number": _cell_str(ws, row, ClaimCol.ACCOUNT_NUMBER),
    }


def has_claim_id(ws, row: int) -> bool:
    """Check if a row already has a Claim ID (i.e., already processed)."""
    val = ws.cell(row=row, column=ClaimCol.CLAIM_ID).value
    if val is None:
        return False
    val_str = str(val).strip()
    # Consider it processed if it has a non-empty value that isn't an error marker
    return bool(val_str) and not val_str.startswith("ERROR:")
